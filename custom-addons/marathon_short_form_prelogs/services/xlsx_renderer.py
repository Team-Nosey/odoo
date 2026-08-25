# -*- coding: utf-8 -*-

import io
import re
from copy import copy
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from odoo import tools

_TEMPLATE_PATH = (
    "marathon_short_form_prelogs/static/src/xlsx/short_form_prelog.xlsx"
)
_SHEET_NAME = "Sheet1"
_FIRST_DATA_ROW = 7
_ORIGINAL_BLANK_ROW = 8
_ORIGINAL_DISCLAIMER_ROW = 9
_LAST_DATA_COLUMN = 12


def render_short_form_workbook(recipient):
    """Render one untouched template copy for one contact."""
    with tools.file_open(_TEMPLATE_PATH, mode="rb") as template_file:
        workbook = load_workbook(io.BytesIO(template_file.read()))

    worksheet = workbook[_SHEET_NAME]
    prelogs = sorted(recipient.prelog_ids, key=_prelog_sort_key)
    if not prelogs:
        raise ValueError("No prelog rows were supplied to the workbook renderer.")

    worksheet["D1"] = recipient.network_id.display_name
    worksheet["D2"] = recipient.week

    worksheet.unmerge_cells(
        start_row=_ORIGINAL_DISCLAIMER_ROW,
        start_column=1,
        end_row=_ORIGINAL_DISCLAIMER_ROW,
        end_column=9,
    )
    worksheet.delete_rows(_ORIGINAL_BLANK_ROW, 1)

    additional_rows = len(prelogs) - 1
    if additional_rows:
        worksheet.insert_rows(_ORIGINAL_BLANK_ROW, additional_rows)

    for row_number in range(_FIRST_DATA_ROW, _FIRST_DATA_ROW + len(prelogs)):
        if row_number != _FIRST_DATA_ROW:
            _copy_row_style(worksheet, _FIRST_DATA_ROW, row_number)

    for row_number, prelog in enumerate(prelogs, start=_FIRST_DATA_ROW):
        deal = prelog.schedule.deal_parent
        values = (
            recipient.network_id.display_name,
            prelog.agency or "",
            prelog.advertiserproduct or "",
            _format_date(prelog.airdate),
            prelog.scheduletime or "",
            prelog.schedulelength or "",
            prelog.orderproductdescription or "",
            prelog.rate or 0.0,
            prelog.scheduleadid or "",
            prelog.name or "",
            deal.estimate or "",
            deal.access_code or "",
        )
        for column_number, value in enumerate(values, start=1):
            worksheet.cell(row=row_number, column=column_number, value=value)

    disclaimer_row = _FIRST_DATA_ROW + len(prelogs)
    worksheet.merge_cells(
        start_row=disclaimer_row,
        start_column=1,
        end_row=disclaimer_row,
        end_column=9,
    )
    _update_table_ranges(worksheet, disclaimer_row)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.read(), _build_filename(recipient)


def _copy_row_style(worksheet, source_row, target_row):
    worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[source_row].height
    for column_number in range(1, _LAST_DATA_COLUMN + 1):
        source = worksheet.cell(row=source_row, column=column_number)
        target = worksheet.cell(row=target_row, column=column_number)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)


def _update_table_ranges(worksheet, last_data_row):
    for table in worksheet.tables.values():
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        if min_row <= _FIRST_DATA_ROW <= max_row:
            table.ref = (
                f"{worksheet.cell(min_row, min_col).coordinate}:"
                f"{worksheet.cell(last_data_row - 1, max_col).coordinate}"
            )


def _prelog_sort_key(prelog):
    return (
        prelog.airdate or datetime.min.date(),
        _time_sort_value(prelog.scheduletime),
        prelog.id,
    )


def _time_sort_value(value):
    if not value:
        return (99, 99, 99)
    text = str(value).strip().upper().replace(" ", "")
    if text.endswith("A") and not text.endswith("AM"):
        text += "M"
    elif text.endswith("P") and not text.endswith("PM"):
        text += "M"
    for time_format in ("%H:%M:%S", "%H:%M", "%I:%M:%S%p", "%I:%M%p"):
        try:
            parsed = datetime.strptime(text, time_format)
            return (parsed.hour, parsed.minute, parsed.second)
        except ValueError:
            continue
    return (98, 98, 98)


def _format_date(value):
    return value.strftime("%m/%d/%Y") if value else ""


def _build_filename(recipient):
    account_name = (
        recipient.account_id.display_name
        or recipient.contact_id.display_name
        or "Contact"
    )
    raw_name = (
        f"{account_name} - {recipient.network_id.display_name} Prelog - "
        f"Week of {recipient.week.isoformat()}.xlsx"
    )
    sanitized = re.sub(r"[\\/:*?\"<>|]+", "-", raw_name)
    sanitized = re.sub(r"\s+", " ", sanitized).strip(" .")
    return sanitized
