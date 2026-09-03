# -*- coding: utf-8 -*-
"""Phase 25 - Bundle Paperwork wizard + Related-tab exposure.

Flow:
  1. Deal form has a "Bundle Paperwork" button. Clicking it opens a
     small modal (mv.bundle_paperwork.wizard) that carries a single
     Generate Paperwork primary button.
  2. Generate builds one XML per station (Salesforce XMLBundleTegnaV2)
     and one Excel summary, saves each as an ir.attachment linked to
     the deal, AND appends each attachment to a dedicated Many2many
     on the deal so the Related tab picks them up.
  3. The Related tab shows the "Bundle Paperwork Files" section
     because the M2M is stored (see the RELATED_TAB_CONFIG addition
     in static/src/js/related_tab/mv_related_tab.js).
"""
import base64
import io
import logging
import re
import zipfile
from collections import OrderedDict
from datetime import date, timedelta

from odoo import models, fields, api, _
from odoo.exceptions import UserError

# Canonical broadcast-quarter maths lives in phase 12 (used by the
# Units + Capping reports). Reuse it so every surface agrees on where
# a quarter begins and ends.
from .phase12_deal_start_date import _broadcast_quarter_bounds

_logger = logging.getLogger(__name__)

_DAY_NAMES = [
    'Monday', 'Tuesday', 'Wednesday',
    'Thursday', 'Friday', 'Saturday', 'Sunday',
]
_GROUPS_TO_LIMIT = {'group_5', 'group_6', 'group_7', 'group_8'}


# ---------------------------------------------------------------------
# Small utilities
# ---------------------------------------------------------------------
def _pad2(n):
    try:
        n = int(n)
    except (TypeError, ValueError):
        return '00'
    return '%02d' % n


def _iso_date(d):
    if not d:
        return ''
    return '%d-%s-%s' % (d.year, _pad2(d.month), _pad2(d.day))


def _string_formatter(text):
    if text is None:
        return ''
    return str(text).replace('&', 'and')


def _iso_week_monday(d):
    """Monday of the week containing `d`."""
    if not d:
        return d
    return d - timedelta(days=d.weekday())


def _quarter_mondays(ref):
    """Every Monday of the BROADCAST quarter containing `ref`.

    Only Mondays whose full Mon..Sun week stays inside the quarter are
    included, so the list is 13 entries (14 in a 53-week broadcast
    year). Example for 2026-08-31 -> broadcast Q3 2026 runs
    2026-06-29..2026-09-27, giving Mondays 6-29 through 9-21.

    Delegates the quarter maths to phase 12 so Bundle Paperwork, the
    Units Report and the Capping Report all agree on where a quarter
    starts and ends.
    """
    if not ref:
        return []
    q_start, q_end = _broadcast_quarter_bounds(_iso_week_monday(ref))
    mondays = []
    cur = q_start
    while cur + timedelta(days=6) <= q_end:
        mondays.append(cur)
        cur += timedelta(days=7)
    return mondays


def _sched_units(schedule):
    """Effective unit count for a schedule, as a float.

    Prefers `units_aired` (booked minus preempted, 0 when canceled).
    Falls back to `units_available` when units_aired is empty, which
    covers rows whose stored compute predates the units_aired fix and
    has not been recomputed yet. A canceled schedule always yields 0.

    Without this fallback the Excel's Sheet2 column M (Total Units)
    resolves to 0 for every schedule, which is what made the Order
    sheet's SUMIF-per-week cells all render as 0.
    """
    if not schedule:
        return 0.0
    if getattr(schedule, 'status', '') == 'canceled':
        return 0.0
    aired = float(getattr(schedule, 'units_aired', 0.0) or 0.0)
    if aired:
        return aired
    available = float(getattr(schedule, 'units_available', 0.0) or 0.0)
    preempted = float(getattr(schedule, 'units_preempted', 0.0) or 0.0)
    return max(0.0, available - preempted)


def _bundle_action_slug(label):
    """Turn a Bundle Action label into a filename-safe token.

    'ADD TO SCHEDULE'                    -> 'Add_To_Schedule'
    'FREQUENCY ADJUSTMENT/CANCEL'        -> 'Frequency_Adjustment_Cancel'
    'FREQUENCY REVISION/TRAFFIC CHANGE'  -> 'Frequency_Revision_Traffic_Change'

    Keeps the token readable in a filename while stripping the
    characters Windows / Odoo attachment names dislike ('/', '\\', ':').
    """
    if not label:
        return 'New_Buy'
    text = str(label).strip()
    # Title-case each alphabetic word, leave digits alone.
    text = re.sub(r'[^0-9A-Za-z]+', ' ', text).strip()
    if not text:
        return 'New_Buy'
    parts = [p.capitalize() if p.isalpha() else p for p in text.split()]
    return '_'.join(parts)


def _civilian_to_military(time_code):
    if not time_code or not isinstance(time_code, str) or not time_code.startswith('v_'):
        return '00:00'
    body = time_code[2:]
    if len(body) < 6:
        return '00:00'
    try:
        hh = int(body[0:2])
        mm = int(body[3:5])
    except ValueError:
        return '00:00'
    suf = body[5]
    if suf == 'a':
        if hh == 12:
            return '00:%s' % _pad2(mm)
        return '%s:%s' % (_pad2(hh), _pad2(mm))
    if suf == 'p':
        if hh == 12:
            return '12:%s' % _pad2(mm)
        return '%s:%s' % (_pad2(hh + 12), _pad2(mm))
    return '00:00'


def _time_label(time_code, time_map):
    if not time_code:
        return ''
    return time_map.get(time_code) or time_code


class _ScheduleChunk:
    __slots__ = (
        'chunk_start', 'chunk_end', 'week_count',
        'total_units', 'bpu_total_units', 'schedules',
    )

    def __init__(self, start_wk, end_wk, count, total, bpu_total, scheds):
        self.chunk_start = start_wk
        self.chunk_end = end_wk
        self.week_count = count
        self.total_units = int(total or 0)
        self.bpu_total_units = int(bpu_total or 0)
        self.schedules = list(scheds)


class _StationGroup:
    __slots__ = ('call_letters', 'market_name', 'pricing_list')

    def __init__(self, call_letters, market_name):
        self.call_letters = call_letters or ''
        self.market_name = market_name or ''
        self.pricing_list = []


# =====================================================================
# ir.attachment - two tag columns for filtering
# =====================================================================
class IrAttachmentBundlePaperwork(models.Model):
    _inherit = 'ir.attachment'

    mv_bundle_paperwork_kind = fields.Selection(
        selection=[('xml', 'XML'), ('excel', 'Excel')],
        string='Bundle Paperwork Kind',
        index=True,
    )
    mv_bundle_paperwork_station = fields.Char(
        string='Bundle Paperwork Station',
    )


# =====================================================================
# mv.deal - button entry point + Related-tab M2M
# =====================================================================
class MvDealBundlePaperwork(models.Model):
    _inherit = 'mv.deal'

    # Stored Many2many so the Related tab's _find_direct_relation
    # picks this up as a real O2M/M2M field on mv.deal. The wizard's
    # generate action populates it after creating the attachments.
    bundle_paperwork_attachment_ids = fields.Many2many(
        'ir.attachment',
        relation='mv_deal_bundle_paperwork_rel',
        column1='deal_id',
        column2='attachment_id',
        string='Notes & Attachments',
    )

    def action_open_bundle_paperwork(self):
        """Return an ir.actions.client so the OWL Bundle Paperwork
        Dialog (registered under the tag below) opens with this
        deal's id in scope. The dialog then calls the RPC methods on
        this model to read state + trigger actions."""
        self.ensure_one()
        return {
            'type': 'ir.actions.client',
            'tag': 'mv_bundle_paperwork_dialog',
            'params': {'deal_id': self.id},
        }

    # ================================================================
    # RPC surface for the OWL dialog
    # ================================================================
    def _mv_bundle_action_label(self, action_code=None):
        """Resolve a Bundle Action code to its human label.

        Passing action_code=None uses the value currently stored on
        the deal. Returns '' when nothing is set so callers can apply
        their own fallback.
        """
        self.ensure_one()
        code = action_code if action_code is not None else (self.bundle_action or '')
        if not code:
            return ''
        return dict(
            self._fields['bundle_action'].selection or []
        ).get(code, code)

    def bundle_paperwork_state(self):
        """Snapshot the dialog needs to render:
          * bundle info (name, brand)
          * bundle action selection options
          * current bundle_action / bundle_start_week values on this Deal
          * generated files (xml + excel) with their attachment ids
        Returns a JSON-friendly dict.
        """
        self.ensure_one()
        Attachment = self.env['ir.attachment']
        atts = Attachment.search([
            ('res_model', '=', 'mv.deal'),
            ('res_id', '=', self.id),
            ('mv_bundle_paperwork_kind', 'in', ['xml', 'excel']),
        ], order='create_date desc, id desc')

        def _serialize(att):
            return {
                'id': att.id,
                'name': att.name,
                'kind': att.mv_bundle_paperwork_kind or '',
                'station': att.mv_bundle_paperwork_station or '',
                'create_date': fields.Datetime.to_string(att.create_date) or '',
            }

        action_selection = self._fields['bundle_action'].selection or []
        bundle_actions = [
            {'code': code, 'label': label}
            for code, label in action_selection
        ]
        return {
            'deal_id': self.id,
            'deal_name': self.name or '',
            'brand': self.brands.name if self.brands else '',
            'program': self.program.display_name if self.program else '',
            'bundle_action': self.bundle_action or '',
            'bundle_start_week': (
                self.bundle_start_week.isoformat()
                if self.bundle_start_week else ''
            ),
            'bundle_actions': bundle_actions,
            'xml_files':   [_serialize(a) for a in atts if a.mv_bundle_paperwork_kind == 'xml'],
            'excel_files': [_serialize(a) for a in atts if a.mv_bundle_paperwork_kind == 'excel'],
        }

    def bundle_paperwork_generate(self, bundle_start_week=False):
        """Delegate to the wizard for the actual XML + Excel build,
        then return the fresh state so the OWL dialog can re-render
        without a page refresh.

        `bundle_start_week` is the week the user picked in the dialog's
        prompt. When supplied it is written to the Deal first, so the
        generated paperwork (and the broadcast quarter its week columns
        span) is anchored on the operator's choice rather than on a
        stale stored value.
        """
        self.ensure_one()
        if bundle_start_week:
            self.write({'bundle_start_week': bundle_start_week})
        wizard = self.env['mv.bundle_paperwork.wizard'].create({
            'deal_id': self.id,
        })
        wizard.action_generate_paperwork()
        return self.bundle_paperwork_state()

    def bundle_paperwork_send(self, attachment_id):
        """Return the act_window that opens the mail composer with
        the file pre-attached. Frontend does_action's it."""
        self.ensure_one()
        att = self.env['ir.attachment'].browse(int(attachment_id)).exists()
        if not att:
            raise UserError(_("Attachment not found."))
        return {
            'type': 'ir.actions.act_window',
            'name': _('Send %s') % (att.name or ''),
            'res_model': 'mail.compose.message',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'default_model': 'mv.deal',
                'default_res_ids': [self.id],
                'default_composition_mode': 'comment',
                'default_attachment_ids': [(6, 0, [att.id])],
                'default_subject': att.name or '',
            },
        }

    def bundle_paperwork_regenerate(self, attachment_id):
        """Regenerate the station XML archive in place. Rewrites the
        same ir.attachment so the download url stays valid. Returns the
        fresh state.

        Station XMLs are packaged as a single zip, so a regenerate
        rebuilds every station and repacks the whole archive rather
        than touching one member.
        """
        self.ensure_one()
        att = self.env['ir.attachment'].browse(int(attachment_id)).exists()
        if not att:
            raise UserError(_("Attachment not found."))
        if att.mv_bundle_paperwork_kind != 'xml':
            raise UserError(_(
                "Regenerate is only available for the XML archive."
            ))
        wizard = self.env['mv.bundle_paperwork.wizard'].new({
            'deal_id': self.id,
        })
        wizard._validate_deal(self)
        chunks = wizard._build_schedule_chunks(self)
        stations = wizard._group_bundle_pricing_by_station(self)
        if not stations:
            raise UserError(_(
                "This program has no active Bundle Pricing records, so "
                "no station XML can be produced."
            ))
        action_label = self._mv_bundle_action_label() or 'NEW BUY'
        today_stamp = _iso_date(fields.Date.context_today(self))
        contact_acc = (self.contactaccount or '').replace('/', '')
        brand_name = ((self.brands and self.brands.name) or '').replace('/', '')

        zip_buf = io.BytesIO()
        with zipfile.ZipFile(
            zip_buf, 'w', compression=zipfile.ZIP_DEFLATED,
        ) as zf:
            for call, station in stations.items():
                body = wizard._generate_xml_for_station(
                    self, chunks, station, self.name or '',
                )
                member_name = '%s-%s-%s-%s-%s-%s.xml' % (
                    call, self.name or '', contact_acc,
                    brand_name, action_label, today_stamp,
                )
                zf.writestr(member_name, body)
        zip_buf.seek(0)
        att.write({
            'datas': base64.b64encode(zip_buf.getvalue()),
            'mimetype': 'application/zip',
        })
        return self.bundle_paperwork_state()

    def bundle_paperwork_run_action(self, action_code, bundle_start_week):
        """Save the picked Bundle Action + Bundle Start Week on this
        Deal, then trigger the corresponding downstream logic. The
        placeholder here just writes the two fields and posts a
        chatter note; wire the real workflow behind this method as
        the specific action-per-workflow gets defined."""
        self.ensure_one()
        # Validate the action_code against the Selection options.
        valid_codes = {c for c, _l in (self._fields['bundle_action'].selection or [])}
        if action_code and action_code not in valid_codes:
            raise UserError(_("Unknown Bundle Action: %s") % action_code)
        vals = {}
        if action_code:
            vals['bundle_action'] = action_code
        if bundle_start_week:
            vals['bundle_start_week'] = bundle_start_week
        if vals:
            self.write(vals)
        # Chatter log so the action is auditable even before the
        # per-workflow logic is filled in.
        action_label = dict(
            self._fields['bundle_action'].selection or []
        ).get(action_code, action_code or '')
        # Update the EXISTING Excel attachment (don't create a new one).
        # If somehow no Excel exists yet (user hit a bundle action
        # before Generate Paperwork), fall through to a fresh generate
        # so the flow doesn't dead-end.
        Attachment = self.env['ir.attachment']
        excel_atts = Attachment.search([
            ('res_model', '=', 'mv.deal'),
            ('res_id', '=', self.id),
            ('mv_bundle_paperwork_kind', '=', 'excel'),
        ], order='create_date desc, id desc')
        if excel_atts:
            self._bundle_paperwork_update_excel(
                excel_atts[0], action_label, bundle_start_week,
            )
        else:
            self.bundle_paperwork_generate()
            # After generate, tag the new Excel with the current action.
            excel_atts = Attachment.search([
                ('res_model', '=', 'mv.deal'),
                ('res_id', '=', self.id),
                ('mv_bundle_paperwork_kind', '=', 'excel'),
            ], order='create_date desc, id desc', limit=1)
            if excel_atts:
                self._bundle_paperwork_update_excel(
                    excel_atts[0], action_label, bundle_start_week,
                )
        # Chatter note.
        if hasattr(self, 'message_post'):
            try:
                self.message_post(body=_(
                    "Bundle Action <b>%(action)s</b> triggered with "
                    "Start Week <b>%(week)s</b>. Existing Excel updated."
                ) % {
                    'action': action_label,
                    'week': bundle_start_week or '',
                })
            except Exception:
                _logger.exception(
                    "bundle_paperwork_run_action: chatter post failed "
                    "for deal id=%s", self.id,
                )
        return self.bundle_paperwork_state()

    def _bundle_paperwork_update_excel(self, attachment, action_label, week):
        """Open the given Excel attachment, rewrite the bundle-action
        label everywhere it appears (row 3 col E on Order + every
        Action cell in the group data rows), and save the bytes back
        to the SAME attachment. Filename is renamed to reflect the
        current action while keeping the deal / brand identifier."""
        import base64
        import io
        from openpyxl import load_workbook
        try:
            raw = base64.b64decode(attachment.datas or b'')
            if not raw:
                return
            wb = load_workbook(io.BytesIO(raw))
        except Exception:
            _logger.exception(
                "bundle_paperwork_update_excel: could not load "
                "existing Excel attachment id=%s", attachment.id,
            )
            return
        if 'Order' not in wb.sheetnames:
            return
        order = wb['Order']
        # Row 3 col E = action label. Everywhere else where the
        # previous action label appears in col B (the Action column
        # inside each group's data rows) gets rewritten too.
        try:
            prev = order['E3'].value or ''
        except Exception:
            prev = ''
        order['E3'] = action_label
        if prev and prev != action_label:
            for row in order.iter_rows(min_col=2, max_col=2):
                cell = row[0]
                if cell.value == prev:
                    cell.value = action_label
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        # Rename to reflect the new action but keep the rest of the
        # filename intact. Generated names look like
        #   <deal>-<contactaccount>-<brand>-<Action_Slug>-<date>.xlsx
        # so we swap the slug segment in place. If the name predates
        # the action-slug convention (or was renamed by hand), the
        # slug is inserted just before the trailing date instead.
        new_name = self._mv_rename_with_action(
            attachment.name or '', action_label,
        )
        attachment.write({
            'datas': base64.b64encode(buf.getvalue()),
            'name': new_name or attachment.name,
        })

    def _mv_rename_with_action(self, filename, action_label):
        """Return `filename` with its Bundle Action token set to
        `action_label`'s slug.

        Handles three shapes:
          1. Name already carries a known action slug  -> replace it.
          2. Name carries a legacy UPPERCASE label     -> replace it.
          3. Name carries neither                      -> insert the
             slug immediately before the trailing date segment.
        """
        if not filename:
            return filename
        new_slug = _bundle_action_slug(action_label)
        stem, dot, ext = filename.rpartition('.')
        if not dot:
            stem, ext = filename, ''
        # Every possible slug / legacy label for this model's selection.
        selection = self._fields['bundle_action'].selection or []
        known_labels = [label for _code, label in selection]
        known_slugs = [_bundle_action_slug(label) for label in known_labels]

        parts = stem.split('-')
        # Case 1 + 2: a segment matches a known slug or legacy label.
        for idx, seg in enumerate(parts):
            if seg in known_slugs or seg in known_labels:
                parts[idx] = new_slug
                break
        else:
            # Case 3: insert before a trailing YYYY-MM-DD date if the
            # tail looks like one, else append at the end.
            if len(parts) >= 3 and re.fullmatch(r'\d{4}', parts[-3] or ''):
                parts.insert(-3, new_slug)
            else:
                parts.append(new_slug)
        rebuilt = '-'.join(p for p in parts if p != '')
        return '%s.%s' % (rebuilt, ext) if ext else rebuilt


# =====================================================================
# Wizard - small modal shell with Generate Paperwork button
# =====================================================================
class MvBundlePaperworkWizard(models.TransientModel):
    _name = 'mv.bundle_paperwork.wizard'
    _description = 'Bundle Paperwork Wizard'

    deal_id = fields.Many2one(
        'mv.deal', string='Deal', required=True, ondelete='cascade',
    )

    # Vestigial column from a prior schema. Kept as optional Char so
    # Odoo's _auto_init drops any leftover NOT NULL constraint on
    # databases that carry the old shape.
    state = fields.Char(default='ready')

    attachment_ids = fields.Many2many(
        'ir.attachment',
        compute='_compute_attachments',
        string='Generated Files',
    )

    @api.depends('deal_id')
    def _compute_attachments(self):
        for rec in self:
            rec.attachment_ids = rec.deal_id.bundle_paperwork_attachment_ids \
                if rec.deal_id else False

    def action_generate_paperwork(self):
        """Build the XML + Excel attachments and link them to the deal."""
        self.ensure_one()
        deal = self.deal_id
        if not deal:
            raise UserError(_("No Deal selected."))
        self._validate_deal(deal)

        Attachment = self.env['ir.attachment']

        # Wipe any prior bundle paperwork on this deal (both physical
        # attachments and the M2M linkage - unlink cascades to the rel).
        existing = Attachment.search([
            ('res_model', '=', 'mv.deal'),
            ('res_id', '=', deal.id),
            ('mv_bundle_paperwork_kind', 'in', ['xml', 'excel']),
        ])
        if existing:
            existing.unlink()

        chunks = self._build_schedule_chunks(deal)
        stations = self._group_bundle_pricing_by_station(deal)
        today_stamp = _iso_date(fields.Date.context_today(self))
        contact_acc = (deal.contactaccount or '').replace('/', '')
        brand_name = ((deal.brands and deal.brands.name) or '').replace('/', '')

        created_ids = []

        # Current Bundle Action label - drives both the in-file action
        # marker and the generated file names. Falls back to NEW BUY so
        # a deal with no action selected keeps the legacy behaviour.
        action_label = deal._mv_bundle_action_label() or 'NEW BUY'
        action_slug = _bundle_action_slug(action_label)

        # ---- one XML per station, collected into ONE zip ---------
        # Individual station XMLs are NOT attached to the deal any
        # more; they only exist as members of the zip archive.
        xml_members = []
        for call, station in stations.items():
            body = self._generate_xml_for_station(
                deal, chunks, station, deal.name or '',
            )
            filename = '%s-%s-%s-%s-%s-%s.xml' % (
                call, deal.name or '', contact_acc,
                brand_name, action_label, today_stamp,
            )
            xml_members.append((filename, body))

        if xml_members:
            zip_buf = io.BytesIO()
            with zipfile.ZipFile(
                zip_buf, 'w', compression=zipfile.ZIP_DEFLATED,
            ) as zf:
                for member_name, member_body in xml_members:
                    zf.writestr(member_name, member_body)
            zip_buf.seek(0)
            zip_name = '%s-%s-%s-%s-%s.zip' % (
                deal.name or 'DEAL', contact_acc, brand_name,
                action_slug, today_stamp,
            )
            att = Attachment.create({
                'name': zip_name,
                'datas': base64.b64encode(zip_buf.getvalue()),
                'res_model': 'mv.deal',
                'res_id': deal.id,
                'mimetype': 'application/zip',
                'type': 'binary',
                'mv_bundle_paperwork_kind': 'xml',
                # Station is intentionally blank: the archive spans
                # every station rather than one.
                'mv_bundle_paperwork_station': '',
            })
            created_ids.append(att.id)

        # ---- one Excel summary --------------------------------
        xlsx_bytes = self._generate_excel(
            deal, chunks, stations, action_label=action_label,
        )
        xlsx_name = '%s-%s-%s-%s-%s.xlsx' % (
            deal.name or 'DEAL', contact_acc, brand_name,
            action_slug, today_stamp,
        )
        att = Attachment.create({
            'name': xlsx_name,
            'datas': base64.b64encode(xlsx_bytes),
            'res_model': 'mv.deal',
            'res_id': deal.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'type': 'binary',
            'mv_bundle_paperwork_kind': 'excel',
        })
        created_ids.append(att.id)

        # ---- link the new set into the deal's Related M2M ----
        # Replace (6, 0, ids) so the deal only points to the *current*
        # generation. Previous ids are already unlinked above; this
        # keeps the M2M rel table in sync.
        deal.write({
            'bundle_paperwork_attachment_ids': [(6, 0, created_ids)],
        })

        # ---- chatter log so the message timeline also shows them ----
        if hasattr(deal, 'message_post'):
            try:
                names = ', '.join(Attachment.browse(created_ids).mapped('name'))
                deal.message_post(
                    body=_("Bundle Paperwork generated: %s") % names,
                    attachment_ids=created_ids,
                )
            except Exception:
                _logger.exception(
                    "Bundle Paperwork: message_post failed for deal id=%s",
                    deal.id,
                )

        self.invalidate_recordset(['attachment_ids'])
        # Reopen the wizard so the file list refreshes in place.
        return {
            'type': 'ir.actions.act_window',
            'name': _('Bundle Paperwork'),
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
            'context': self.env.context,
        }

    # ================================================================
    # Validation
    # ================================================================
    def _validate_deal(self, deal):
        missing = []
        if not deal.advertiser:
            missing.append(_('Advertiser'))
        if not (deal.brands and deal.brands.name):
            missing.append(_('Brand name'))
        # Bundle Start Week is intentionally NOT required: when it is
        # blank (or points at a week with no Schedule) the chunk builder
        # falls back to the current broadcast quarter's Schedules and
        # back-fills the field with the week it actually used. See
        # _resolve_bundle_schedules.
        if not deal.contactaccount:
            missing.append(_('Contact Account'))
        if not deal.length:
            missing.append(_('Length'))
        if not deal.program:
            missing.append(_('Program'))
        if missing:
            raise UserError(_(
                "The following field(s) are required to generate "
                "bundle paperwork: %s"
            ) % ', '.join(missing))

    # ================================================================
    # Schedule chunk builder
    # ================================================================
    def _resolve_bundle_schedules(self, deal):
        """Pick the schedule set to build paperwork from.

        Resolution order:

        1. If the Deal has a Bundle Start Week AND a Schedule exists on
           exactly that week, use everything from that week forward.
           (Original behaviour - an explicit start wins.)
        2. Otherwise fall back to the CURRENT BROADCAST quarter: use the
           Schedules whose week falls inside the quarter that contains
           today. This is what makes "Generate Paperwork" work on a
           Deal whose Bundle Start Week is blank or points at a week
           with no Schedule.
        3. If the quarter has no Schedules either, use the earliest
           remaining un-sent Schedule from today onward, so paperwork
           can still be produced for a purely future flight.

        Returns (schedules_recordset, effective_start_week).
        """
        Schedules = self.env['mv.schedules']
        today = fields.Date.context_today(self)

        base_domain = [
            ('deal_parent', '=', deal.id),
            ('xml_sent', '=', False),
        ]

        # ---- 1. explicit Bundle Start Week, exact match ------------
        if deal.bundle_start_week:
            exact = Schedules.search(
                base_domain + [('week', '=', deal.bundle_start_week)],
                limit=1,
            )
            if exact:
                schedules = Schedules.search(
                    base_domain + [('week', '>=', deal.bundle_start_week)],
                    order='week asc',
                )
                return schedules, deal.bundle_start_week

        # ---- 2. current broadcast quarter --------------------------
        q_start, q_end = _broadcast_quarter_bounds(today)
        if q_start and q_end:
            schedules = Schedules.search(
                base_domain + [
                    ('week', '>=', q_start),
                    ('week', '<=', q_end),
                ],
                order='week asc',
            )
            if schedules:
                _logger.info(
                    "Bundle Paperwork: deal id=%s has no Schedule on its "
                    "Bundle Start Week (%s); falling back to the current "
                    "broadcast quarter %s..%s (%s schedule(s)).",
                    deal.id, deal.bundle_start_week, q_start, q_end,
                    len(schedules),
                )
                return schedules, schedules[0].week

        # ---- 3. earliest future un-sent schedule -------------------
        schedules = Schedules.search(
            base_domain + [('week', '>=', _iso_week_monday(today))],
            order='week asc',
        )
        if schedules:
            _logger.info(
                "Bundle Paperwork: deal id=%s - no Schedule in the current "
                "broadcast quarter; using the earliest future Schedule week %s.",
                deal.id, schedules[0].week,
            )
            return schedules, schedules[0].week

        return Schedules.browse(), None

    def _build_schedule_chunks(self, deal):
        schedules, effective_start = self._resolve_bundle_schedules(deal)
        if not schedules:
            q_start, q_end = _broadcast_quarter_bounds(
                fields.Date.context_today(self)
            )
            raise UserError(_(
                "No usable Schedules were found on this Deal.\n\n"
                "Checked:\n"
                "  * the Deal's Bundle Start Week (%(bsw)s)\n"
                "  * the current broadcast quarter (%(qs)s to %(qe)s)\n"
                "  * any later un-sent Schedule\n\n"
                "Make sure this Deal has at least one Schedule with "
                "XML Sent unchecked."
            ) % {
                'bsw': deal.bundle_start_week or _('not set'),
                'qs': q_start,
                'qe': q_end,
            })

        # Keep the Deal's Bundle Start Week in step with what we
        # actually used, so the Excel header block, the mini-popup
        # default and any later regenerate all agree.
        if effective_start and deal.bundle_start_week != effective_start:
            deal.sudo().write({'bundle_start_week': effective_start})

        chunks = []
        cur = _ScheduleChunk(
            schedules[0].week,
            schedules[0].week + timedelta(days=6),
            1,
            int(_sched_units(schedules[0])),
            int(schedules[0].bpu_units or 0),
            [schedules[0]],
        )
        for s in schedules[1:]:
            if s.week != cur.chunk_end + timedelta(days=1):
                chunks.append(cur)
                cur = _ScheduleChunk(
                    s.week, s.week + timedelta(days=6), 1,
                    int(_sched_units(s)),
                    int(s.bpu_units or 0),
                    [s],
                )
            else:
                cur.chunk_end = cur.chunk_end + timedelta(days=7)
                cur.week_count += 1
                cur.total_units += int(_sched_units(s))
                cur.bpu_total_units += int(s.bpu_units or 0)
                cur.schedules.append(s)
        chunks.append(cur)
        return chunks

    # ================================================================
    # Station grouping
    # ================================================================
    def _group_bundle_pricing_by_station(self, deal):
        BundlePricing = self.env['mv.bundle_pricing']
        pricing = BundlePricing.search([
            ('program', '=', deal.program.id),
            ('active', '=', True),
        ])
        if not pricing:
            raise UserError(_(
                "No active Bundle Pricing records were found for this "
                "Deal's Program. Add at least one active Bundle Pricing."
            ))
        stations = OrderedDict()
        for bp in pricing:
            call = bp.call_letters or ''
            if not call:
                continue
            if call not in stations:
                market_name = (
                    bp.station.market.name
                    if bp.station and bp.station.market else ''
                )
                stations[call] = _StationGroup(call, market_name)
            stations[call].pricing_list.append(bp)
        if not stations:
            raise UserError(_(
                "None of the active Bundle Pricing records have a Station "
                "with Call Letters."
            ))
        return stations

    # ================================================================
    # XML per station
    # ================================================================
    def _generate_xml_for_station(self, deal, chunks, station, order_id):
        now = fields.Datetime.now()
        timestamp = now.strftime('%Y-%m-%dT%H:%M:%S.000Z')
        deal_account = (deal.contactaccount or '')[:34]

        buyline_body, order_total = self._generate_buylines(
            deal, chunks, station,
        )
        chunk_start = _iso_date(chunks[0].chunk_start)
        chunk_end   = _iso_date(chunks[-1].chunk_end)

        parts = []
        parts.append('<?xml version="1.0" encoding="UTF-8"?>\n')
        parts.append(
            '<tvo:CreateOrderRequest serviceName="CreateOrder" '
            'originatingTradingPartner="00000000-0000-0000-0000-000000000000" '
            'messageExpiration="%s" timestamp="%s" messageVersion="3.2" '
            'destinationTradingPartner="00000000-0000-0000-0000-000000000000" '
            'serviceInstanceId="00000000-0000-0000-0000-000000000000" '
            'mediaType="SpotTV" messageId="00000000-0000-0000-0000-000000000000" '
            'serviceVersion="3.2" targetEnvironment="PRODUCTION" '
            'isRetransmission="False" '
            'd1p1:schemaLocation="http://www.tvb.org/schema/TVB_OrderCommon TVB_CreateOrderRequest_3.2.xsd" '
            'xmlns:tvo="http://www.tvb.org/schema/TVB_OrderCommon" '
            'xmlns:tvc="http://www.tvb.org/schema/TVB_Common" '
            'xmlns:tvm="http://www.tvb.org/schema/TVB_MediaCommon" '
            'xmlns:d1p1="http://www.w3.org/2001/XMLSchema-instance">\n\n'
            % (timestamp, timestamp)
        )
        parts.append(
            '<tvo:Order orderVersion="1" orderStatus="New" orderId="%s">\n'
            '<tvo:OrderIdReferences>\n'
            '<tvc:SourceCode source="Agency">National</tvc:SourceCode>\n'
            '<tvc:SourceCode source="Rep"></tvc:SourceCode>\n'
            '</tvo:OrderIdReferences>\n\n'
            % order_id
        )
        parts.append(
            '<tvo:AltOrderIdReferences>\n'
            '<tvc:SourceCode source="Agency"></tvc:SourceCode>\n'
            '</tvo:AltOrderIdReferences>\n\n'
        )
        parts.append(
            '<tvo:OrderType>Normal</tvo:OrderType>\n'
            '<tvo:OrderCashTrade>Cash</tvo:OrderCashTrade>\n\n'
        )
        parts.append(
            '<tvm:Advertiser>\n'
            '<tvc:CompanyName>%s(M)</tvc:CompanyName>\n'
            '<tvc:SourceCode source="Agency"></tvc:SourceCode>\n'
            '</tvm:Advertiser>\n\n'
            % _string_formatter(deal.advertiser)
        )
        parts.append(
            '<tvm:Product>\n'
            '<tvm:ProductName>%s</tvm:ProductName>\n'
            '<tvc:SourceCode source="Agency"></tvc:SourceCode>\n'
            '</tvm:Product>\n\n'
            % _string_formatter(deal.brands.name if deal.brands else '')
        )
        parts.append(
            '<tvm:Agency>\n'
            '<tvc:CompanyName>Marathon Ventures (Comm)</tvc:CompanyName>\n'
            '<tvc:Address addressRole="Billing">\n'
            '<tvc:Street1>675 Third Avenue, 11th Floor</tvc:Street1>\n'
            '<tvc:City>New York</tvc:City>\n'
            '<tvc:RegionCode>NY</tvc:RegionCode>\n'
            '<tvc:PostalCode>10017</tvc:PostalCode>\n'
            '<tvc:CountryCode></tvc:CountryCode>\n'
            '</tvc:Address>\n\n'
            '<tvc:Contact contactRole="Buyer">\n'
            '<tvc:PersonFirstName>Tania</tvc:PersonFirstName>\n'
            '<tvc:PersonLastName>Bonetti</tvc:PersonLastName>\n'
            '<tvc:Email></tvc:Email>\n'
            '<tvc:Phone></tvc:Phone>\n'
            '</tvc:Contact>\n\n'
            '<tvc:SourceCode source="Station">00000000-0000-0000-0000-000000000000</tvc:SourceCode>\n'
            '<tvc:Office>\n'
            '<tvc:OfficeName>%s c/o Marathon</tvc:OfficeName>\n'
            '<tvc:SourceCode source="Agency">00000000-0000-0000-0000-000000000000</tvc:SourceCode>\n'
            '</tvc:Office>\n'
            '</tvm:Agency>\n\n'
            % deal_account
        )
        parts.append(
            '<tvm:Estimate>\n'
            '<tvc:SourceCode source="Agency">%s EST:%s</tvc:SourceCode>\n'
            '</tvm:Estimate>\n\n'
            % (deal_account, order_id)
        )

        logs_first = ''
        logs_last = ''
        if deal.program and deal.program.logs_contact:
            full_name = deal.program.logs_contact.name or ''
            pieces = full_name.split(' ', 1)
            logs_first = pieces[0] if pieces else ''
            logs_last  = pieces[1] if len(pieces) > 1 else ''
        bundle_name = (deal.program.bundle_name if deal.program else '') or ''
        parts.append(
            '<tvm:Seller>\n'
            '<tvm:StationSeller>\n'
            '<tvc:CompanyName>%s - %s - %s</tvc:CompanyName>\n'
            '<tvc:Contact contactRole="AccountExec">\n'
            '<tvc:PersonFirstName>%s</tvc:PersonFirstName>\n'
            '<tvc:PersonLastName>%s</tvc:PersonLastName>\n'
            '<tvc:Email></tvc:Email>\n'
            '<tvc:Phone></tvc:Phone>\n'
            '</tvc:Contact>\n'
            '<tvc:SourceCode source="Station">00000000-0000-0000-0000-000000000000</tvc:SourceCode>\n'
            '<tvc:Office>\n'
            '<tvc:OfficeName>Marathon</tvc:OfficeName>\n'
            '<tvc:SourceCode source="Agency">00000000-0000-0000-0000-000000000000</tvc:SourceCode>\n'
            '</tvc:Office>\n'
            '</tvm:StationSeller>\n'
            '</tvm:Seller>\n\n'
            % (station.call_letters, station.market_name, bundle_name,
               logs_first, logs_last)
        )
        parts.append(
            '<tvo:LocalNational>National</tvo:LocalNational>\n'
            '<tvm:Station>\n'
            '<tvm:FCCCallLetters>%s</tvm:FCCCallLetters>\n'
            '<tvc:SourceCode source="Station">00000000-0000-0000-0000-000000000000</tvc:SourceCode>\n'
            '</tvm:Station>\n\n'
            % station.call_letters
        )
        parts.append(
            '<tvc:StartDate>%s</tvc:StartDate>\n'
            '<tvc:EndDate>%s</tvc:EndDate>\n'
            '<tvo:OrderGrossAmount>%.2f</tvo:OrderGrossAmount>\n'
            '<tvo:BillingCalendar>Broadcast</tvo:BillingCalendar>\n'
            '<tvo:BillingCycle>Monthly</tvo:BillingCycle>\n'
            '<tvm:PrimaryDemoCategory demoId="DM0" />\n'
            '<tvm:DemoCategory demoId="DM0">\n'
            '<tvm:DemoGroup>Adults</tvm:DemoGroup>\n'
            '<tvm:DemoLowerAge>18</tvm:DemoLowerAge>\n'
            '<tvm:DemoUpperAge>99</tvm:DemoUpperAge>\n'
            '</tvm:DemoCategory>\n'
            '<tvc:Comment source="Agency"></tvc:Comment>\n\n'
            % (chunk_start, chunk_end, order_total)
        )
        parts.append(buyline_body)
        parts.append('</tvo:Order>\n</tvo:CreateOrderRequest>\n')
        return ''.join(parts)

    # ================================================================
    # Buylines
    # ================================================================
    def _generate_buylines(self, deal, chunks, station):
        order_total = 0.0
        parts = ['<tvo:Buylines>\n\n']
        count = 1
        length_num = (deal.length or '').lstrip('v_') or '30'
        try:
            length_int = int(length_num)
        except ValueError:
            length_int = 30
        program_name = (deal.program.name if deal.program else '') or ''

        for chunk in chunks:
            for bp in station.pricing_list:
                units = int(bp.units or 0) * chunk.total_units
                if bp.group in _GROUPS_TO_LIMIT and \
                   program_name == 'Primary Tegna Connect':
                    units = int(bp.units or 0) * chunk.bpu_total_units
                rate = round(
                    (float(bp.rate_per_30 or 0.0) * length_int) / 30, 2,
                )
                total = round(units * rate, 2)
                order_total += total

                parts.append(
                    '<tvo:SpotBuyline buylineNumber="%d" buylineVersion="1" '
                    'buylineStatus="New">\n'
                    '<tvo:BuylineIdReferences>\n'
                    '<tvc:SourceCode source="Agency"></tvc:SourceCode>\n'
                    '</tvo:BuylineIdReferences>\n\n'
                    % count
                )
                parts.append(
                    '<tvo:BuylineCashTrade>Cash</tvo:BuylineCashTrade>\n'
                    '<tvo:BuylineDescription>Sign-On/Sign-Off</tvo:BuylineDescription>\n'
                    '<tvc:StartDate>%s</tvc:StartDate>\n'
                    '<tvc:EndDate>%s</tvc:EndDate>\n'
                    '<tvo:BuylineQuantity unitType="Spot">%d</tvo:BuylineQuantity>\n'
                    '<tvo:BuylineUnitRate costModel="Unit">%.2f</tvo:BuylineUnitRate>\n'
                    '<tvo:BuylineGrossAmount>%.2f</tvo:BuylineGrossAmount>\n'
                    '<tvc:Comment source="Agency"></tvc:Comment>\n'
                    '<tvo:SpotBuylineType>Normal</tvo:SpotBuylineType>\n'
                    '<tvo:SpotLength>%d</tvo:SpotLength>\n'
                    '<tvm:StartDayOfWeek>Mo</tvm:StartDayOfWeek>\n\n'
                    % (_iso_date(chunk.chunk_start),
                       _iso_date(chunk.chunk_end),
                       units, rate, total, length_int)
                )
                parts.append('<tvo:ContractInterval>\n')
                bp_day_codes = set(
                    (bp.days.mapped('name') or [])
                    + (bp.days.mapped('code') or [])
                )
                days_block_parts = []
                for day in _DAY_NAMES:
                    valid = 'True' if day in bp_day_codes else 'False'
                    days_block_parts.append(
                        '<tvm:%sValid>%s</tvm:%sValid>\n' % (day, valid, day)
                    )
                days_block = ''.join(days_block_parts)
                parts.append(days_block)
                parts.append(
                    '<tvc:StartTime>%s</tvc:StartTime>\n'
                    '<tvc:EndTime>%s</tvc:EndTime>\n'
                    '</tvo:ContractInterval>\n\n'
                    % (_civilian_to_military(bp.start_time),
                       _civilian_to_military(bp.end_time))
                )
                parts.append(self._weekly_spot_distribution(
                    days_block, chunk, int(bp.units or 0),
                    bp.group or '', program_name,
                ))
                parts.append('</tvo:SpotBuyline>\n\n')
                count += 1
        parts.append('</tvo:Buylines>\n')
        return ''.join(parts), round(order_total, 2)

    def _weekly_spot_distribution(self, days_block, chunk,
                                  bp_units, group, program_name):
        parts = []
        for sc in chunk.schedules:
            units = int(_sched_units(sc) * bp_units)
            if group in _GROUPS_TO_LIMIT and program_name == 'Primary Tegna Connect':
                units = int((sc.bpu_units or 0) * bp_units)
            end_wk = sc.week + timedelta(days=6) if sc.week else sc.week
            parts.append(
                '<tvo:WeeklySpotDistribution>\n'
                '<tvc:StartDate>%s</tvc:StartDate>\n'
                '<tvc:EndDate>%s</tvc:EndDate>\n'
                '<tvo:SpotPerWeekQuantity>%d</tvo:SpotPerWeekQuantity>\n'
                % (_iso_date(sc.week), _iso_date(end_wk), units)
            )
            parts.append(days_block)
            parts.append('</tvo:WeeklySpotDistribution>\n')
        return ''.join(parts)

    # ================================================================
    # Excel summary - built programmatically from scratch.
    #
    # Seven sheets, laid out in the traffic-team's format:
    #
    #   * Order        - header block + one Group section per group,
    #                    with the SUMIF / VLOOKUP / SUMPRODUCT formulas
    #                    referencing Sheet2 + Station List.
    #   * Sheet2       - one row per (schedule, week).
    #   * Station List - one row per active Bundle Pricing record,
    #                    with a CONCAT lookup key in column A.
    #   * Schedule     - schedule metadata row(s).
    #   * Sheet1       - blank scratch sheet.
    #   * Quarter Set  - broadcast-quarter lookup (Monday -> quarter
    #                    start). 13-week quarters starting 2016-03-28,
    #                    covering ~10 years and going 5 years forward
    #                    from today.
    #   * Stations     - unique station call letters.
    # ================================================================
    def _generate_excel(self, deal, chunks, stations, action_label='NEW BUY'):
        from datetime import date, timedelta
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        time_map = dict(
            self.env['mv.bundle_pricing']._fields['start_time'].selection or []
        )
        group_selection = (
            self.env['mv.bundle_pricing']._fields['group'].selection or []
        )
        group_label_map = dict(group_selection)

        # ================================================================
        # Pre-computed lookups so the Order sheet can be SELF-CONTAINED.
        #
        # Requirement: the delivered workbook must contain ONLY the
        # "Order" tab. Three Order cells used to be cross-sheet
        # formulas, which would resolve to #REF! the moment their source
        # sheet is removed:
        #
        #   A  No. of Stations  <- COUNTIF('Station List'!B, <group>)
        #   G  Station Rate     <- SUMIF('Station List'!E, C<row>, ..G)
        #   K..X Weekly units   <- SUMIF(Sheet2!A, <monday>, Sheet2!M)
        #
        # We therefore evaluate them here, in Python, against the same
        # data the helper sheets were built from, and write literal
        # values into Order. Intra-sheet formulas (H, I, J, the totals)
        # are untouched - they only reference Order itself, so they keep
        # recalculating live in Excel.
        # ================================================================

        # 1. Weekly units per Monday - mirrors Sheet2 column M, which is
        #    =SUM(K,L) = units_aired + bonus units, summed per week.
        weekly_units_by_monday = {}
        for ch in chunks:
            for sc in ch.schedules:
                if not sc.week:
                    continue
                total = _sched_units(sc) + float(sc.bpu_units or 0)
                weekly_units_by_monday[sc.week] = (
                    weekly_units_by_monday.get(sc.week, 0.0) + total
                )

        # 2. Sum of Rate per :30 keyed by Time Period text - mirrors
        #    SUMIF('Station List'!E, <period>, 'Station List'!G).
        # 3. Count of Station List rows per Group label - mirrors
        #    COUNTIF('Station List'!B, <group label>).
        rate_by_period = {}
        rows_by_group_label = {}
        for _call, station in stations.items():
            for bp in station.pricing_list:
                period_txt = '%s-%s' % (
                    _time_label(bp.start_time, time_map),
                    _time_label(bp.end_time, time_map),
                )
                rate_by_period[period_txt] = (
                    rate_by_period.get(period_txt, 0.0)
                    + float(bp.rate_per_30 or 0.0)
                )
                label = group_label_map.get(bp.group, bp.group or '')
                rows_by_group_label[label] = (
                    rows_by_group_label.get(label, 0) + 1
                )

        # Group codes present on this deal, in the order the group
        # selection defines them (Group 1, Group 2, ...). Group them
        # once - we need both the list of groups AND the list of
        # distinct time periods per group (so the Order sheet emits
        # exactly N data rows per group, matching the sample where
        # each station carries 3 rows in Station List, one per period).
        groups_present = []          # [(group_code, group_label)]
        periods_per_group = {}       # group_code -> list of (start_end_str, telecast_days)
        for group_code, group_label in group_selection:
            seen = []
            for st in stations.values():
                for bp in st.pricing_list:
                    if bp.group != group_code:
                        continue
                    period = '%s-%s' % (
                        _time_label(bp.start_time, time_map),
                        _time_label(bp.end_time, time_map),
                    )
                    telecast = 'MTWTFSS'
                    if hasattr(bp, 'days') and bp.days:
                        # 'MTWTFSS'-like short label; fall back to full week
                        try:
                            names = list(bp.days.mapped('name') or [])
                            initials = {'Mon': 'M', 'Tue': 'T', 'Wed': 'W',
                                        'Thu': 'T', 'Fri': 'F',
                                        'Sat': 'S', 'Sun': 'S'}
                            telecast = ''.join(initials.get(n, '') for n in names) or 'MTWTFSS'
                        except Exception:
                            telecast = 'MTWTFSS'
                    entry = (period, telecast)
                    if entry not in seen:
                        seen.append(entry)
            if seen:
                groups_present.append((group_code, group_label))
                periods_per_group[group_code] = seen
        if not groups_present:
            groups_present = [('group_1', 'Group 1')]
            periods_per_group['group_1'] = [
                ('09:00A-12:00A', 'MTWTFSS'),
                ('12:00A-04:00A', 'MTWTFSS'),
                ('04:00A-09:00A', 'MTWTFSS'),
            ]

        program = deal.program
        bundle_name = (program.bundle_name if program else '') or (
            program.name if program else ''
        )
        # Bundle "family" title shown in row 3 col A - Gray / Tegna /
        # Hearst / etc. Pulled off program.name; fall back to bundle_name.
        family = ''
        prog_name = (program.name if program else '') or ''
        upper = prog_name.upper()
        for kw in ('GRAY', 'TEGNA', 'HEARST', 'UNIVISION', 'UNIMAS',
                   'AMERICAN SPIRIT', 'BOUNCE', 'PRIMARY'):
            if kw in upper:
                family = kw
                break
        family = family or (bundle_name or '').upper() or 'BUNDLE'

        logs_contact = program.logs_contact if program else None
        logs_name = (logs_contact.name if logs_contact else '') or ''
        logs_phone = (
            logs_contact.phone
            if logs_contact and 'phone' in logs_contact._fields else ''
        ) or ''
        logs_email = (
            logs_contact.email
            if logs_contact and 'email' in logs_contact._fields else ''
        ) or ''

        wb = Workbook()
        # Sheet order matches the sample: Order, Schedule, Sheet2,
        # Station List, RB Market List, Quarter Set.
        order = wb.active
        order.title = 'Order'
        sch = wb.create_sheet('Schedule')
        sh2 = wb.create_sheet('Sheet2')
        sl = wb.create_sheet('Station List')
        rb = wb.create_sheet('RB Market List')
        qs = wb.create_sheet('Quarter Set')

        # -------------------- styles --------------------
        bold = Font(bold=True)
        heading_font = Font(bold=True, size=12)
        title_font = Font(bold=True, size=14)
        header_fill = PatternFill(
            start_color='D9D9D9', end_color='D9D9D9', fill_type='solid',
        )
        center = Alignment(horizontal='center', vertical='center')

        # ================================================================
        # Order sheet
        # ================================================================
        # ---- Header block (rows 3-11) ---------------------------------
        # Row 3: family (col A) + bundle-action label (col E).
        order['A3'] = family
        order['A3'].font = title_font
        order['E3'] = action_label
        order['E3'].font = title_font

        # Row 5: Network / Name
        order['A5'] = 'Network:';         order['A5'].font = bold
        order['B5'] = bundle_name
        order['L5'] = 'Name:';            order['L5'].font = bold
        order['M5'] = logs_name

        # Row 6: No. of Stations (formula) / Phone
        order['A6'] = 'No. of Stations:'; order['A6'].font = bold
        # Sum the group No.-of-Stations cells from each block. Populated
        # once we know the block start rows further down.
        order['L6'] = 'Phone:';           order['L6'].font = bold
        order['M6'] = logs_phone

        # Row 7: Agency / Email
        order['A7'] = 'Agency:';          order['A7'].font = bold
        order['B7'] = deal.contactaccount or ''
        order['L7'] = 'Email:';           order['L7'].font = bold
        order['M7'] = logs_email

        # Row 8: Advertiser / Date
        order['A8'] = 'Advertiser:';      order['A8'].font = bold
        order['B8'] = deal.advertiser or ''
        order['L8'] = 'Date:';            order['L8'].font = bold
        order['M8'] = _iso_date(fields.Date.context_today(self))

        # Row 9: Product
        order['A9'] = 'Product:';         order['A9'].font = bold
        order['B9'] = deal.brands.name if deal.brands else ''

        # Row 10: Access Code
        order['A10'] = 'Access Code:';    order['A10'].font = bold
        order['B10'] = deal.access_code or ''

        # Row 11: Deal
        order['A11'] = 'Deal:';           order['A11'].font = bold
        order['B11'] = deal.name or ''

        # ---- Group blocks -------------------------------------------
        # Layout PER GROUP (matches the sample workbook):
        #   +0 heading (col A: "GROUP N", uppercase)
        #   +1 column headers (No. of Stations, Action, Time Period, ...)
        #        + K-X: 14 successive-Monday formulas
        #        (no Z/AA helper columns - K..X read Sheet2 directly)
        #   +2 .. +2+N-1 : one data row per distinct time period in
        #        this group's Bundle Pricing
        #   +N+2 blank
        #   +N+3 Comments (col B), Station Total (col I,J)
        #   +N+4 Group Total (col I,J)
        # Blocks are separated by one blank row.
        length_num = (deal.length or '').lstrip('v_') or '30'
        try:
            length_int = int(length_num)
        except ValueError:
            length_int = 30

        block_no_of_stations_cells = []   # 'A15', 'A24', ... for row 6 SUM
        cur_row = 13
        for group_code, group_label in groups_present:
            periods = periods_per_group.get(group_code) or [
                ('09:00A-12:00A', 'MTWTFSS'),
            ]
            n_periods = len(periods)

            # No. of Stations for this group, evaluated up-front because
            # the Station Rate (col G) divides by it. Mirrors the old
            # ROUND(COUNTIF('Station List'!B,<group>)/n_periods,0).
            _grp_rows = int(rows_by_group_label.get(group_label, 0))
            _stations_for_group = (
                int(round(_grp_rows / n_periods)) if n_periods else 0
            )

            r_head       = cur_row                # "GROUP N"
            r_columns    = cur_row + 1            # column headers
            r_data_first = cur_row + 2            # first data row
            r_data_last  = r_data_first + n_periods - 1  # last data row
            r_blank      = r_data_last + 1
            r_totals_stn = r_blank + 1
            r_totals_grp = r_totals_stn + 1

            # Heading uppercased ("GROUP 1")
            order.cell(row=r_head, column=1,
                       value=group_label.upper()).font = heading_font

            # Column headers row
            col_headers = [
                'No. of Stations', 'Action', 'Time Period', 'Telecast Days',
                'Len', 'Max/Day', 'Station Rate', 'Group Rate',
                'No. Wks', 'No. Units',
            ]
            for idx, hdr in enumerate(col_headers, start=1):
                cell = order.cell(row=r_columns, column=idx, value=hdr)
                cell.font = bold
                cell.fill = header_fill
                cell.alignment = center
            # K..X = the week columns. These span the BROADCAST QUARTER
            # that the paperwork belongs to - every Monday whose full
            # Mon..Sun week fits inside the quarter - NOT an arbitrary
            # 14 weeks counted forward from the first schedule.
            #
            # e.g. generating on 2026-08-31 puts us in broadcast Q3 2026
            # (2026-06-29..2026-09-27), so the columns run
            # 6-29-2026 ... 9-21-2026 (13 of them). 9-28 is excluded
            # because its week spills into Q4.
            #
            # Each header is written as a real date (not a `=prev+7`
            # formula) so the quarter boundary is exact and the trailing
            # columns can be left genuinely empty.
            quarter_ref = (
                deal.bundle_start_week
                or (chunks[0].chunk_start if chunks else None)
                or fields.Date.context_today(self)
            )
            quarter_mondays = _quarter_mondays(quarter_ref)
            max_week_cols = 14                     # K(11)..X(24)
            week_mondays = quarter_mondays[:max_week_cols]
            week_col_count = len(week_mondays)
            for offset, monday in enumerate(week_mondays):
                cell = order.cell(
                    row=r_columns, column=11 + offset, value=monday,
                )
                cell.font = bold
                cell.number_format = 'm/d/yyyy'
            # Blank any unused slot so a short quarter doesn't leave a
            # stale header behind.
            for offset in range(week_col_count, max_week_cols):
                order.cell(row=r_columns, column=11 + offset, value=None)
            # Data rows - one per distinct time period. K-X show the
            # raw units-per-week straight from Sheet2 (no station-tier
            # multiplier); Station Rate is the per-30 rate scaled to
            # the row's length. Every value formula is wrapped in
            # IFERROR so an empty Station List / missing schedule
            # data renders as a blank cell instead of #DIV/0! / #N/A.
            for i, (period, telecast) in enumerate(periods):
                data_row = r_data_first + i
                order.cell(row=data_row, column=2, value=action_label)
                order.cell(row=data_row, column=3, value=period)
                order.cell(row=data_row, column=4, value=telecast)
                order.cell(row=data_row, column=5, value=length_int)
                # G: Station Rate = SUMIF(rate) * (E/30) / A<stations>
                # Uses the row's Time Period text to match the Station
                # List's Start/End Time column directly (no Concat/
                # Multiplier round-trip).
                # Literal: (sum of Rate per :30 for this Time Period)
                #          * (length / 30) / No. of Stations
                # Replaces the SUMIF against the Station List sheet.
                _period_rate = float(rate_by_period.get(period, 0.0))
                _station_rate = 0.0
                if _stations_for_group:
                    _station_rate = (
                        _period_rate * (float(length_int or 0) / 30.0)
                    ) / _stations_for_group
                order.cell(
                    row=data_row, column=7,
                    value=round(_station_rate, 2),
                )
                # H: Group Rate = G * <No. of Stations>
                order.cell(
                    row=data_row, column=8,
                    value='=IFERROR(G%d*A%d,0)' % (data_row, r_data_first),
                )
                order.cell(
                    row=data_row, column=9,
                    value='=IFERROR(COUNTIF(K%d:X%d,">0"),0)' % (
                        data_row, data_row,
                    ),
                )
                order.cell(
                    row=data_row, column=10,
                    value='=IFERROR(SUM(K%d:X%d),0)' % (data_row, data_row),
                )
                # K..X: raw units-per-week from Sheet2. No multiplier.
                # Only emit a formula for columns that actually carry a
                # week header - a 13-week quarter leaves X empty, and a
                # SUMIF against a blank key would render a stray 0.
                # Literal per-week totals, looked up by that column's
                # Monday. Replaces the SUMIF against Sheet2.
                for offset in range(week_col_count):
                    col = 11 + offset
                    monday = week_mondays[offset]
                    order.cell(
                        row=data_row, column=col,
                        value=weekly_units_by_monday.get(monday, 0),
                    )

            # Col A first-data-row = No. of Stations. IFERROR keeps
            # divide-by-zero out of the header SUM when the Station
            # List is empty.
            # Literal (computed above) instead of a COUNTIF against the
            # Station List sheet, so it survives that sheet's removal.
            order.cell(
                row=r_data_first, column=1,
                value=_stations_for_group,
            ).font = bold
            block_no_of_stations_cells.append('A%d' % r_data_first)

            # Comments (col B) + Station Total (col I,J)
            order.cell(row=r_totals_stn, column=2, value='Comments:').font = bold
            order.cell(row=r_totals_stn, column=9, value='Station Total').font = bold
            order.cell(
                row=r_totals_stn, column=10,
                value='=IFERROR(SUMPRODUCT(G%d:G%d,J%d:J%d),0)' % (
                    r_data_first, r_data_last, r_data_first, r_data_last,
                ),
            )
            order.cell(row=r_totals_grp, column=9, value='Group Total').font = bold
            order.cell(
                row=r_totals_grp, column=10,
                value='=IFERROR(A%d*J%d,0)' % (r_data_first, r_totals_stn),
            )

            cur_row = r_totals_grp + 2   # one blank row between blocks

        # Row 6 col B: No. of Stations = SUM(A15, A24, ...)
        if block_no_of_stations_cells:
            order['B6'] = '=SUM(%s)' % ','.join(block_no_of_stations_cells)
        else:
            order['B6'] = len(stations)

        # ---- Footer: "*All times ..." + Traffic Instructions --------
        r_note = cur_row
        order.cell(row=r_note, column=1,
                   value='*All times noted above are in Eastern Time').font = bold

        r_traffic_hdr = r_note + 1
        order.cell(row=r_traffic_hdr, column=1,
                   value='Traffic Instructions:').font = bold

        r_columns_traf = r_traffic_hdr + 1
        # Column headers - Flight Week (A), Toll Free # (B), Spot 1 (C),
        # Spot 2 (I). Uses the same 8-col grouping as the sample.
        order.cell(row=r_columns_traf, column=1, value='Flight Week').font = bold
        order.cell(row=r_columns_traf, column=2, value='Toll Free #').font = bold
        order.cell(row=r_columns_traf, column=3, value='Spot 1').font = bold
        order.cell(row=r_columns_traf, column=9, value='Spot 2').font = bold

        toll_free = ''
        if program and hasattr(program, 'toll_free'):
            toll_free = program.toll_free or ''

        r_traf = r_columns_traf + 1
        for ch in chunks:
            for sc in ch.schedules:
                week_iso = sc.week.strftime('%m/%d/%Y') if sc.week else ''
                order.cell(row=r_traf, column=1, value=week_iso)
                order.cell(row=r_traf, column=2, value=toll_free)
                # Spot 1: parse "ISCI1/ISCI2" from AE column - split on slash.
                order.cell(
                    row=r_traf, column=3,
                    value=(
                        '=IF(AE%d="","",IF(ISERROR(SEARCH("/",AE%d)),AE%d,'
                        'LEFT(AE%d,SEARCH("/",AE%d)-1)))'
                        % (r_traf, r_traf, r_traf, r_traf, r_traf)
                    ),
                )
                # Spot 2: right side of slash if present.
                order.cell(
                    row=r_traf, column=9,
                    value=(
                        '=IF(ISERROR(SEARCH("/",AE%d)),"",'
                        'MID(AE%d,SEARCH("/",AE%d)+1,'
                        'LEN(AE%d)-SEARCH("/",AE%d)))'
                        % (r_traf, r_traf, r_traf, r_traf, r_traf)
                    ),
                )
                # ISCI raw value on col AE (col 31).
                order.cell(row=r_traf, column=31, value=sc.isci_code or '')
                r_traf += 1

        # Column widths
        for col_letter, width in zip(
            ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'L', 'M'],
            [18, 20, 14, 14, 6, 8, 12, 12, 8, 10, 8, 30],
        ):
            order.column_dimensions[col_letter].width = width
        for wk_col in range(11, 25):
            order.column_dimensions[get_column_letter(wk_col)].width = 11

        # ================================================================
        # Sheet2 (per-schedule-week)
        # ================================================================
        sh2_headers = [
            'Week', 'ContactAccount', 'DealBrand', 'Bundle Units',
            'Bundle Rate', 'Total Dollars', 'Schedule: Schedule Name',
            '', '', '', 'Units Aired', 'Bonus Units', 'Total Units',
        ]
        for c, h in enumerate(sh2_headers, start=1):
            cell = sh2.cell(row=1, column=c, value=h)
            cell.font = bold
            cell.fill = header_fill
        r = 2
        for ch in chunks:
            for sc in ch.schedules:
                total_dollars = float(
                    (sc.units_available or 0.0) * (sc.rate or 0.0)
                )
                sh2.cell(row=r, column=1, value=sc.week)
                sh2.cell(row=r, column=2, value=deal.contactaccount or '')
                sh2.cell(row=r, column=3,
                         value=deal.brands.name if deal.brands else '')
                sh2.cell(row=r, column=4, value=int(sc.units_available or 0))
                sh2.cell(row=r, column=5, value=float(sc.rate or 0.0))
                sh2.cell(row=r, column=6, value=total_dollars)
                sh2.cell(row=r, column=7, value=sc.name or '')
                # K = aired units (falls back to available-minus-preempted
                # so a stale stored compute can't zero out the row), L =
                # bonus units, M = total consumed by the Order sheet's
                # per-week SUMIF.
                sh2.cell(row=r, column=11, value=_sched_units(sc))
                sh2.cell(row=r, column=12, value=int(sc.bpu_units or 0))
                sh2.cell(row=r, column=13, value='=SUM(K%d,L%d)' % (r, r))
                r += 1

        # ================================================================
        # Station List (used by Order sheet SUMIF / VLOOKUP formulas)
        # ================================================================
        sl_headers = [
            'CONCAT', 'Group', 'Market', 'Call Letters',
            'Start/End Time', 'Units', 'Rate per :30', 'Station Rank',
        ]
        for c, h in enumerate(sl_headers, start=1):
            cell = sl.cell(row=1, column=c, value=h)
            cell.font = bold
            cell.fill = header_fill
        r = 2
        for call, station in stations.items():
            for bp in station.pricing_list:
                time_period = '%s-%s' % (
                    _time_label(bp.start_time, time_map),
                    _time_label(bp.end_time, time_map),
                )
                group_label = group_label_map.get(bp.group, bp.group or '')
                sl.cell(row=r, column=1, value='=CONCATENATE(B%d,E%d)' % (r, r))
                sl.cell(row=r, column=2, value=group_label)
                sl.cell(row=r, column=3, value=station.market_name)
                sl.cell(row=r, column=4, value=call)
                sl.cell(row=r, column=5, value=time_period)
                sl.cell(row=r, column=6, value=int(bp.units or 0))
                sl.cell(row=r, column=7, value=float(bp.rate_per_30 or 0.0))
                # Station Rank (H) - not modeled in Odoo; left blank.
                r += 1

        # ================================================================
        # Schedule sheet (metadata) - 14 cols, includes Status
        # ================================================================
        sch_headers = [
            'ContactAccount', 'DealBrand', 'Status', 'Week', 'Days Allowed',
            'UnitLength', 'Units Available', 'Rate', 'Total Dollars',
            'Schedule: Schedule Name', 'Start Time', 'End Time',
            'StartEnd', 'Max/Day',
        ]
        for c, h in enumerate(sch_headers, start=1):
            cell = sch.cell(row=1, column=c, value=h)
            cell.font = bold
            cell.fill = header_fill
        status_map = dict(
            self.env['mv.schedules']._fields['status'].selection or []
        )
        r = 2
        for ch in chunks:
            for sc in ch.schedules:
                total_dollars = float(
                    (sc.units_available or 0.0) * (sc.rate or 0.0)
                )
                days_allowed = ', '.join(
                    sc.days_allowed.mapped('name') or []
                ) if hasattr(sc, 'days_allowed') else ''
                sch.cell(row=r, column=1, value=deal.contactaccount or '')
                sch.cell(row=r, column=2,
                         value=deal.brands.name if deal.brands else '')
                sch.cell(row=r, column=3,
                         value=status_map.get(sc.status, sc.status or 'Sold'))
                sch.cell(row=r, column=4,
                         value=sc.week.strftime('%m/%d/%Y') if sc.week else '')
                sch.cell(row=r, column=5, value=days_allowed)
                sch.cell(row=r, column=6, value=length_int)
                sch.cell(row=r, column=7, value=int(sc.units_available or 0))
                sch.cell(row=r, column=8, value=float(sc.rate or 0.0))
                sch.cell(row=r, column=9, value=total_dollars)
                sch.cell(row=r, column=10, value=sc.name or '')
                sch.cell(row=r, column=11,
                         value=_time_label(sc.start_time, time_map))
                sch.cell(row=r, column=12,
                         value=_time_label(sc.end_time, time_map))
                sch.cell(row=r, column=13,
                         value='%s-%s' % (
                             _time_label(sc.start_time, time_map),
                             _time_label(sc.end_time, time_map),
                         ))
                sch.cell(row=r, column=14, value=int(sc.max_per_day or 0))
                r += 1

        # ================================================================
        # RB Market List - static reference; header row + notice.
        # ================================================================
        rb.cell(row=1, column=2, value='RAYCOM').font = bold
        rb.cell(row=8, column=1, value='RANK').font = bold
        rb.cell(row=8, column=2, value='MARKET').font = bold
        rb.cell(row=8, column=3, value='%US').font = bold
        rb.cell(row=8, column=4, value='PRIMARY STATION').font = bold
        rb.cell(row=8, column=5, value='SECONDARY STATION').font = bold
        rb.cell(row=8, column=6, value='TV HOMES').font = bold
        # Groups + market rows would be populated per-program from a
        # curated reference table; not modeled in Odoo yet, so this
        # sheet ships with only the header structure.

        # ================================================================
        # Quarter Set (13-week broadcast quarters, Monday keys)
        # ================================================================
        qs.cell(row=1, column=1, value='Week').font = bold
        qs.cell(row=1, column=2, value='Quarter Start').font = bold
        anchor = date(2016, 3, 28)
        today = fields.Date.context_today(self)
        end_date = date(today.year + 5, 12, 31)
        weeks = (end_date - anchor).days // 7 + 1
        for i in range(weeks):
            monday = anchor + timedelta(days=7 * i)
            quarter_index = i // 13
            quarter_start = anchor + timedelta(days=7 * 13 * quarter_index)
            qs.cell(row=2 + i, column=1, value=monday)
            qs.cell(row=2 + i, column=2, value=quarter_start)
        qs.column_dimensions['A'].width = 14
        qs.column_dimensions['B'].width = 16

        # ================================================================
        # Deliver ONLY the "Order" tab.
        #
        # The helper sheets above are still built, because the Order
        # sheet's layout is derived from the same data and keeping them
        # makes the two easy to compare while developing. They are
        # dropped here, immediately before saving.
        #
        # This is safe ONLY because every Order cell that used to point
        # at 'Station List' or Sheet2 is now a Python-computed literal
        # (see the pre-computed lookups near the top of this method).
        # If you ever re-introduce a cross-sheet formula in Order, it
        # will become #REF! at this point.
        # ================================================================
        for sheet_name in list(wb.sheetnames):
            if sheet_name != 'Order':
                del wb[sheet_name]

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
